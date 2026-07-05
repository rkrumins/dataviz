"""xlsx FormatAdapter — a real Excel workbook with a **Nodes** sheet + **Edges** sheet, locked
identity columns, an Instructions sheet, and an ``_op`` dropdown.

This is the strategic fix for flat-CSV fragility (a hand-added column or a trimmed trailing cell
silently shifts a value into the wrong column, e.g. a property landing in ``_op``): in Excel each
column is a named header and the kind comes from the SHEET, not a ``kind`` column, so adding a
property means typing under a ``prop.<name>`` header — nothing shifts. Needs openpyxl, so it lives
in its own module and registers lazily in :data:`formats._ADAPTERS` (a missing lib never breaks the
other formats).
"""
from __future__ import annotations

import io
from typing import Any, AsyncIterator, Dict, List, Sequence

# Identity / system columns — greyed in the header so users know not to touch them.
_LOCKED = {"entity_id", "urn", "baseVersion", "source_entity_id", "target_entity_id"}
_NODE_SHEET = "Nodes"
_EDGE_SHEET = "Edges"
_INSTRUCTIONS = "Instructions"


class XlsxAdapter:
    fmt = "xlsx"

    async def parse(self, chunks: AsyncIterator[bytes]) -> AsyncIterator[Dict[str, Any]]:
        import openpyxl  # lazy — only when xlsx is actually used

        buf = b""
        async for chunk in chunks:
            buf += chunk
        wb = openpyxl.load_workbook(io.BytesIO(buf), read_only=True, data_only=True)
        try:
            for sheet_name, kind in ((_NODE_SHEET, "node"), (_EDGE_SHEET, "edge")):
                ws = _find_sheet(wb, sheet_name)
                if ws is None:
                    continue
                for rec in _read_sheet(ws):
                    rec["kind"] = kind          # kind is the SHEET, not a column
                    yield rec
        finally:
            wb.close()

    async def write(
        self, records: AsyncIterator[Dict[str, Any]], *, columns: Sequence[str]
    ) -> AsyncIterator[bytes]:
        import openpyxl

        node_recs: List[Dict[str, Any]] = []
        edge_recs: List[Dict[str, Any]] = []
        async for rec in records:
            (edge_recs if rec.get("kind") == "edge" else node_recs).append(rec)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)                     # drop the default blank sheet
        _write_instructions(wb.create_sheet(_INSTRUCTIONS))
        _write_sheet(wb.create_sheet(_NODE_SHEET), node_recs, columns)
        _write_sheet(wb.create_sheet(_EDGE_SHEET), edge_recs, columns)
        out = io.BytesIO()
        wb.save(out)
        yield out.getvalue()


# --------------------------------------------------------------------------- #
def _find_sheet(wb, name: str):
    for ws in wb.worksheets:
        if str(ws.title or "").strip().lower() == name.lower():
            return ws
    return None


def _read_sheet(ws) -> List[Dict[str, Any]]:
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return []
    header = [str(h).strip() if h is not None else "" for h in header_row]
    out: List[Dict[str, Any]] = []
    for row in rows:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue                             # blank row
        rec: Dict[str, Any] = {}
        for i, val in enumerate(row):
            if i < len(header) and header[i] and val is not None and str(val).strip() != "":
                rec[header[i]] = val             # empty cells dropped (PATCH semantics)
        if rec:
            out.append(rec)
    return out


def _sheet_columns(records: List[Dict[str, Any]], columns: Sequence[str]) -> List[str]:
    """Per-sheet columns: the unified export order, minus ``kind`` (it's the sheet), filtered to
    what this kind actually uses (so edge-only columns don't clutter the Nodes sheet). An empty
    sheet falls back to all non-``kind`` columns so it's still a usable blank template."""
    present = {k for r in records for k in r}
    if not present:
        return [c for c in columns if c != "kind"]
    return [c for c in columns if c != "kind" and c in present]


def _write_sheet(ws, records: List[Dict[str, Any]], columns: Sequence[str]) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    cols = _sheet_columns(records, columns)
    if not cols:
        return
    ws.append(cols)
    header_font = Font(bold=True)
    locked_fill = PatternFill("solid", fgColor="E8E8E8")
    for i, c in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        if c in _LOCKED:
            cell.fill = locked_fill
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(c) + 4))

    for rec in records:
        ws.append([_cell(rec.get(c)) for c in cols])
    ws.freeze_panes = "A2"

    if "_op" in cols and records:
        dv = DataValidation(type="list", formula1='"upsert,delete"', allow_blank=True)
        ws.add_data_validation(dv)
        letter = get_column_letter(cols.index("_op") + 1)
        dv.add(f"{letter}2:{letter}{len(records) + 1}")


def _cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _write_instructions(ws) -> None:
    from openpyxl.styles import Font

    lines = [
        ("How to edit this workbook", 14),
        ("", 0),
        ("• The Nodes and Edges sheets hold your data — one entity per row, and EACH PROPERTY IS ITS"
         " OWN COLUMN (prop.<name>). Edit a cell to change a value.", 0),
        ("• Greyed columns (entity_id, urn, baseVersion, source/target_entity_id) are the identity"
         " — leave them exactly as-is so your edits match the right item.", 0),
        ("• You only need the ROWS and COLUMNS you're changing. Deleting rows or columns from this"
         " file changes nothing — untouched entities and fields are left exactly as they are.", 0),
        ("• Add a NEW property = add a column named 'prop.<name>' (e.g. prop.owner) and fill only the"
         " rows it applies to. Add a new entity = a new row with entity_id/urn left blank.", 0),
        ("• DELETE a property value with the token \\N in its cell (or delete the whole row's entity"
         " by setting _op to 'delete').", 0),
        ("• Save as .xlsx and import — you'll review every change before anything is published.", 0),
    ]
    for text, size in lines:
        ws.append([text])
        if size:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=size)
    ws.column_dimensions["A"].width = 110
